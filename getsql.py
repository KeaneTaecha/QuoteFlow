"""
AUTOMATIC EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION
===============================================================
This script automatically reads your entire Excel file and converts
ALL tables (even multiple tables per sheet) to SQLite database.

Usage:
    python convert_excel_to_sqlite.py

Output:
    Creates 'prices.db' with all your price data
"""

import sqlite3
import openpyxl
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict, Set


@dataclass
class TableLocation:
    """Represents the location and bounds of a detected table"""
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    width_row: int  # Row containing width headers
    height_col: int  # Column containing height headers


class ExcelToSQLiteConverter:
    """Automatically convert Excel price list to SQLite database with multi-table detection"""
    
    def __init__(self, excel_file, db_file='prices.db'):
        self.excel_file = excel_file
        self.db_file = db_file
        self.conn = None
        
        # Will read from Header sheet
        self.header_data = []
        
        # Statistics
        self.stats = {
            'total_sheets': 0,
            'processed_sheets': 0,
            'skipped_sheets': 0,
            'total_prices': 0,
            'total_products': 0,
            'total_tables': 0,
            'errors': []
        }
    
    def create_database(self):
        """Create database structure"""
        print("Creating database structure...")
        
        self.conn = sqlite3.connect(self.db_file)
        cursor = self.conn.cursor()
        
        # Products table
        cursor.execute('''
            CREATE TABLE products (
                product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_id INTEGER NOT NULL,
                model TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                anodized_multiplier REAL,
                powder_coated_multiplier REAL,
                UNIQUE(table_id, model)
            )
        ''')
        
        # Prices table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS prices (
                price_id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_id INTEGER NOT NULL,
                width INTEGER NOT NULL,
                height INTEGER NOT NULL,
                normal_price REAL,
                price_with_damper REAL,
                FOREIGN KEY (table_id) REFERENCES products(table_id) ON UPDATE CASCADE,
                UNIQUE(table_id, width, height)
            )
        ''')
        
        # Indexes for fast lookups
        cursor.execute('''
            CREATE INDEX idx_price_lookup 
            ON prices(table_id, width, height)
        ''')
        
        cursor.execute('''
            CREATE INDEX idx_model_lookup 
            ON products(model)
        ''')
        
        self.conn.commit()
        print("✓ Database structure created\n")
    
    def is_inch_value(self, value) -> Optional[int]:
        """Check if a cell value is an inch measurement (e.g., '4"', '6"')"""
        if value and isinstance(value, str) and '"' in str(value):
            try:
                # Extract inch value from string like "4"" -> 4
                width_str = str(value).replace('"', '').strip()
                return int(width_str)
            except ValueError:
                return None
        return None
    
    def find_table_boundaries(self, sheet, start_row: int, start_col: int, 
                            processed_areas: Set[Tuple[int, int]]) -> Optional[TableLocation]:
        """Find the boundaries of a table starting from a given position"""
        
        # Check if this area was already processed
        if (start_row, start_col) in processed_areas:
            return None
        
        # Look for width row (containing inch values like "4"", "6"")
        # Typically around rows 7-10
        width_row = None
        for row in range(max(1, start_row - 5), min(start_row + 20, sheet.max_row + 1)):
            cell_value = sheet.cell(row, start_col).value
            if self.is_inch_value(cell_value):
                width_row = row
                break
        
        if width_row is None:
            return None
        
        # Find the extent of width columns
        end_col = start_col
        for col in range(start_col, min(start_col + 100, sheet.max_column + 1)):
            cell_value = sheet.cell(width_row, col).value
            if not self.is_inch_value(cell_value):
                end_col = col
                break
        
        # Find height column (should be to the left of the table, containing inch values)
        height_col = None
        for col in range(max(1, start_col - 10), start_col):
            for row in range(width_row + 2, min(width_row + 20, sheet.max_row + 1), 2):
                if self.is_inch_value(sheet.cell(row, col).value):
                    height_col = col
                    break
        
        if height_col is None:
            height_col = 1  # Default to column A
        
        # Find the extent of height rows
        end_row = width_row + 1
        for row in range(width_row + 2, min(sheet.max_row + 1, width_row + 500), 2):
            if not self.is_inch_value(sheet.cell(row, height_col).value):
                break
            end_row = row + 1  # Include the mm row after the inch row
        
        return TableLocation(
            start_row=width_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            width_row=width_row,
            height_col=height_col
        )
    
    def detect_all_tables(self, sheet, expected_count: Optional[int] = None) -> List[TableLocation]:
        """Detect all tables in a sheet using pattern recognition"""
        tables = []
        processed_areas = set()
        max_search_row = min(200, sheet.max_row)  # Limit search depth
        max_search_col = min(100, sheet.max_column)  # Limit search width
        
        # Scan the sheet for table patterns
        for row in range(1, max_search_row + 1):
            for col in range(1, max_search_col + 1):
                # Skip if this cell is already part of a processed table
                if any((row >= t.start_row and row <= t.end_row and 
                       col >= t.start_col and col <= t.end_col) for t in tables):
                    continue
                
                # Look for inch value patterns that might indicate a table
                cell_value = sheet.cell(row, col).value
                if self.is_inch_value(cell_value):
                    # Potential table found, try to determine its boundaries
                    table = self.find_table_boundaries(sheet, row, col, processed_areas)
                    if table:
                        tables.append(table)
                        # Mark this area as processed
                        for r in range(table.start_row, table.end_row + 1):
                            for c in range(table.start_col, table.end_col + 1):
                                processed_areas.add((r, c))
                        
                        # Stop searching if we've found the expected number of tables
                        if expected_count is not None and len(tables) >= expected_count:
                            break
            
            # Break outer loop if we've found the expected number of tables
            if expected_count is not None and len(tables) >= expected_count:
                break
        
        # Sort tables by position (left to right, then top to bottom)
        tables.sort(key=lambda t: (t.start_row, t.start_col))
        
        return tables
    
    def extract_and_store_prices(self, sheet, table_loc: TableLocation, table_id: int) -> int:
        """Extract price data from a table at a specific location and store in database"""
        cursor = self.conn.cursor()
        price_count = 0
        
        # Process each width column
        for col in range(table_loc.start_col, table_loc.end_col + 1):
            width = self.is_inch_value(sheet.cell(table_loc.width_row, col).value)
            if width is None:
                continue
            
            # Process height rows
            for row in range(table_loc.width_row + 2, table_loc.end_row + 1, 2):
                height = self.is_inch_value(sheet.cell(row, table_loc.height_col).value)
                if height is None:
                    continue
                
                # Get prices
                try:
                    # Normal price (inch row)
                    normal_price_cell = sheet.cell(row, col).value
                    normal_price = float(normal_price_cell) if normal_price_cell and isinstance(normal_price_cell, (int, float)) else None
                    
                    # Price with damper (mm row - next row)
                    damper_price_cell = sheet.cell(row + 1, col).value
                    damper_price = float(damper_price_cell) if damper_price_cell and isinstance(damper_price_cell, (int, float)) else None
                    
                    # Insert if at least one price exists
                    if normal_price is not None or damper_price is not None:
                        cursor.execute('''
                            INSERT OR REPLACE INTO prices 
                            (table_id, width, height, normal_price, price_with_damper)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (table_id, width, height, normal_price, damper_price))
                        price_count += 1
                except Exception:
                    continue
        
        return price_count
    
    def read_header_sheet(self, wb):
        """Read the Header sheet to get table metadata"""
        print("Reading Header sheet...")
        
        if 'Header' not in wb.sheetnames:
            print("❌ Error: 'Header' sheet not found in Excel file!")
            return False
        
        header_sheet = wb['Header']
        self.header_data = []
        
        # Group entries by sheet name
        sheet_entries = {}
        
        # Read header data
        for row in header_sheet.iter_rows(min_row=2, values_only=True):
            table_id, sheet_name, model = row[0], row[1], row[2]
            anodized_multiplier, powder_coated_multiplier = row[3], row[4]
            
            if table_id is None or sheet_name is None or model is None:
                continue
            
            # Parse models
            models = [m.strip() for m in str(model).split(',')]
            
            # Parse multipliers
            anodized_multipliers = []
            if anodized_multiplier is not None:
                anodized_multipliers = [float(m.strip()) for m in str(anodized_multiplier).split(',')]
            
            powder_coated_multipliers = []
            if powder_coated_multiplier is not None:
                powder_coated_multipliers = [float(m.strip()) for m in str(powder_coated_multiplier).split(',')]
            
            # Adjust multiplier lists to match model count
            if len(anodized_multipliers) == 1:
                anodized_multipliers = anodized_multipliers * len(models)
            elif len(anodized_multipliers) > 0 and len(anodized_multipliers) < len(models):
                anodized_multipliers.extend([anodized_multipliers[-1]] * (len(models) - len(anodized_multipliers)))
            
            if len(powder_coated_multipliers) == 1:
                powder_coated_multipliers = powder_coated_multipliers * len(models)
            elif len(powder_coated_multipliers) > 0 and len(powder_coated_multipliers) < len(models):
                powder_coated_multipliers.extend([powder_coated_multipliers[-1]] * (len(models) - len(powder_coated_multipliers)))
            
            entry = {
                'table_id': int(table_id),
                'sheet_name': str(sheet_name),
                'models': models,
                'anodized_multipliers': anodized_multipliers,
                'powder_coated_multipliers': powder_coated_multipliers
            }
            
            self.header_data.append(entry)
            
            # Group by sheet name for multi-table detection
            if sheet_name not in sheet_entries:
                sheet_entries[sheet_name] = []
            sheet_entries[sheet_name].append(entry)
        
        # Store grouped entries for easier processing
        self.sheet_entries = sheet_entries
        
        print(f"✓ Found {len(self.header_data)} table(s) in Header sheet")
        print(f"✓ Sheets with multiple tables: {[s for s, e in sheet_entries.items() if len(e) > 1]}\n")
        return True
    
    def insert_products(self, table_id, sheet_name, models, anodized_multipliers, powder_coated_multipliers):
        """Insert product records for a table"""
        cursor = self.conn.cursor()
        
        for i, model in enumerate(models):
            anodized_mult = anodized_multipliers[i] if i < len(anodized_multipliers) else None
            powder_mult = powder_coated_multipliers[i] if i < len(powder_coated_multipliers) else None
            
            cursor.execute('''
                INSERT OR IGNORE INTO products (table_id, model, sheet_name, anodized_multiplier, powder_coated_multiplier)
                VALUES (?, ?, ?, ?, ?)
            ''', (table_id, model, sheet_name, anodized_mult, powder_mult))
            self.stats['total_products'] += 1
    
    def extract_tables_from_sheet(self, sheet, sheet_name):
        """Extract and process all tables from a sheet"""
        entries = self.sheet_entries.get(sheet_name, [])
        if not entries:
            return 0
        
        total_prices = 0
        detected_tables = self.detect_all_tables(sheet, len(entries))
        
        print(f"  Processing {len(entries)} table(s) - detected {len(detected_tables)} table(s)")
        
        # Process each detected table
        for i, table_loc in enumerate(detected_tables):
            if i < len(entries):
                entry = entries[i]
                print(f"    Table {i+1} (ID: {entry['table_id']}, Models: {', '.join(entry['models'][:2])}...)")
                print(f"      Location: Row {table_loc.start_row}, Col {table_loc.start_col}")
                
                # Insert products
                self.insert_products(
                    entry['table_id'], sheet_name, entry['models'],
                    entry['anodized_multipliers'], entry['powder_coated_multipliers']
                )
                
                # Process the table
                price_count = self.extract_and_store_prices(sheet, table_loc, entry['table_id'])
                total_prices += price_count
                self.stats['total_tables'] += 1
                print(f"      ✓ {price_count} prices imported")
            else:
                print(f"    ⚠ Extra table detected at ({table_loc.start_row},{table_loc.start_col}) - no Header entry")
        
        # Warn about missing tables
        if len(entries) > len(detected_tables):
            print(f"    ⚠ Warning: Expected {len(entries)} tables but only found {len(detected_tables)}")
            for j in range(len(detected_tables), len(entries)):
                print(f"      Missing: Table ID {entries[j]['table_id']}")
        
        return total_prices
    
    def convert(self):
        """Main conversion process"""
        print("="*70)
        print("AUTOMATIC EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION")
        print("="*70)
        print(f"\nSource file: {self.excel_file}")
        print(f"Output database: {self.db_file}")
        print()
        
        # Check if Excel file exists
        if not Path(self.excel_file).exists():
            print(f"❌ Error: File '{self.excel_file}' not found!")
            return False
        
        # Create database
        self.create_database()
        
        # Load Excel file
        print(f"Loading Excel file...")
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True, data_only=True)
            print(f"✓ Loaded {len(wb.sheetnames)} sheets\n")
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False
        
        self.stats['total_sheets'] = len(wb.sheetnames)
        
        # Read Header sheet
        if not self.read_header_sheet(wb):
            return False
        
        if not self.header_data:
            print("❌ Error: No table information found in Header sheet!")
            return False
        
        # Process sheets
        print("Processing sheets:")
        print("-" * 70)
        
        processed_sheets = set()
        
        for sheet_name in self.sheet_entries.keys():
            if sheet_name in processed_sheets:
                continue
            
            if sheet_name not in wb.sheetnames:
                print(f"❌ Warning: Sheet '{sheet_name}' not found in Excel file!")
                self.stats['skipped_sheets'] += 1
                continue
            
            print(f"\nProcessing: {sheet_name}")
            
            try:
                sheet = wb[sheet_name]
                price_count = self.extract_tables_from_sheet(sheet, sheet_name)
                
                if price_count > 0:
                    print(f"  Total: {price_count} prices from this sheet")
                    self.stats['processed_sheets'] += 1
                    self.stats['total_prices'] += price_count
                else:
                    print(f"  ⚠ No valid prices found in sheet")
                    self.stats['skipped_sheets'] += 1
                
                processed_sheets.add(sheet_name)
                    
            except Exception as e:
                print(f"  ❌ Error: {str(e)}")
                self.stats['errors'].append(f"{sheet_name}: {str(e)}")
                self.stats['skipped_sheets'] += 1
        
        # Commit all changes
        self.conn.commit()
        self.conn.close()
        
        # Print summary
        self.print_summary()
        
        return True
    
    def print_summary(self):
        """Print conversion summary"""
        print("\n" + "="*70)
        print("CONVERSION COMPLETE!")
        print("="*70)
        print(f"\nTotal sheets in Excel:     {self.stats['total_sheets']}")
        print(f"Sheets processed:          {self.stats['processed_sheets']}")
        print(f"Sheets skipped:            {self.stats['skipped_sheets']}")
        print(f"Total tables found:        {self.stats['total_tables']}")
        print(f"Total products created:    {self.stats['total_products']}")
        print(f"Total prices inserted:     {self.stats['total_prices']:,}")
        
        if self.stats['errors']:
            print(f"\nErrors encountered:        {len(self.stats['errors'])}")
            for error in self.stats['errors'][:5]:
                print(f"  - {error}")
        
        print(f"\n✓ Database saved to: {self.db_file}")
        print(f"✓ File size: {Path(self.db_file).stat().st_size / 1024:.1f} KB")
        print("\nYou can now use this database in your PyQt application!")


# =============================================================================
# MAIN PROGRAM - RUN THIS!
# =============================================================================

if __name__ == "__main__":
    # Configuration
    EXCEL_FILE = 'price_list_modified.xlsx'  # Your Excel file name
    DATABASE_FILE = 'prices.db'               # Output database name
    
    print("\n")
    print("╔════════════════════════════════════════════════════════════════════╗")
    print("║      EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION         ║")
    print("║      Automatically detects and processes multiple tables!         ║")
    print("╚════════════════════════════════════════════════════════════════════╝")
    print("\n")
    
    # Create converter
    converter = ExcelToSQLiteConverter(EXCEL_FILE, DATABASE_FILE)
    
    # Run conversion
    success = converter.convert()
    
    if success:
        print("\n✓ Conversion completed successfully!")
    else:
        print("\n❌ Conversion failed. Please check the errors above.")