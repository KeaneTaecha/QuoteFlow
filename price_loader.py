"""
Price List Loader Module
Handles loading and parsing the Excel price list for HRG and WSG products.
"""

import re
import openpyxl


class PriceListLoader:
    """Loads and parses the price list Excel file"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.price_data = {}
        self.load_prices()
    
    def load_prices(self):
        """Load HRG and WSG prices from the Excel file"""
        products = ['HRG', 'WSG']
        finishes = {'Alu': 'Anodized Aluminum', 'Wh': 'White Powder Coated'}
        
        for product in products:
            self.price_data[product] = {}
            for finish_code, finish_name in finishes.items():
                # The actual sheet names are "1-HRG,2-WSG Alu" and "1-HRG,2-WSG Wh"
                sheet_name = f"1-HRG,2-WSG {finish_code}"
                if sheet_name in self.workbook.sheetnames:
                    # Parse the sheet and get data for the specific product
                    sheet_data = self.parse_sheet(sheet_name, product)
                    self.price_data[product][finish_name] = sheet_data
    
    def parse_sheet(self, sheet_name, product):
        """Parse a price sheet into a structured dictionary for a specific product"""
        sheet = self.workbook[sheet_name]
        prices = {'regular': {}, 'with_damper': {}}
        
        # Dynamically find the width headers range
        start_col = self.find_width_headers_start(sheet)
        if start_col is None:
            return prices  # No width headers found, return empty prices
        
        end_col = self.find_width_headers_end(sheet, start_col)
        if end_col is None:
            return prices  # No end found, return empty prices
        
        # Get width headers from the found range
        width_headers = []
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=8, column=col).value
            if cell_value and isinstance(cell_value, str) and '"' in cell_value:
                width_headers.append(cell_value.strip())
        
        # Find the starting row for the specific product
        start_row = self.find_product_start_row(sheet, product)
        if start_row == -1:
            return prices
        
        # Parse data rows starting from the found row
        row = start_row
        while row < 100:
            # Get height from first column
            height_cell = sheet.cell(row=row, column=1).value
            if not height_cell:
                break
            
            height = str(height_cell).strip()
            
            # Check if this is a regular row (height like "6\"") or WD row (height like 150)
            is_regular = isinstance(height_cell, str) and '"' in height_cell
            is_wd = isinstance(height_cell, (int, float)) and height_cell > 50
            
            # Get prices for each width
            prices_dict = {}
            for col_idx, width in enumerate(width_headers):
                price_cell = sheet.cell(row=row, column=col_idx + start_col).value
                if price_cell and isinstance(price_cell, (int, float)):
                    if is_regular:
                        size_key = f"{width} x {height}"
                        prices_dict[size_key] = float(price_cell)
            
            # Store regular prices
            if is_regular and prices_dict:
                prices['regular'].update(prices_dict)
            
            # Handle WD prices - they are in rows with numeric heights
            if is_wd:
                # Find the previous regular height row to get the size reference
                prev_row = row - 1
                while prev_row >= start_row:
                    prev_height_cell = sheet.cell(row=prev_row, column=1).value
                    if prev_height_cell and isinstance(prev_height_cell, str) and '"' in prev_height_cell:
                        prev_height = str(prev_height_cell).strip()
                        # Get WD prices for each width
                        for col_idx, width in enumerate(width_headers):
                            price_cell = sheet.cell(row=row, column=col_idx + start_col).value
                            if price_cell and isinstance(price_cell, (int, float)):
                                size_key = f"{width} x {prev_height}"
                                prices['with_damper'][size_key] = float(price_cell)
                        break
                    prev_row -= 1
            
            row += 1
        
        return prices
    
    def find_product_start_row(self, sheet, product):
        """Find the starting row for a specific product in the sheet"""
        # Look for product indicators in column 16 (P)
        for row in range(1, 20):
            cell_value = sheet.cell(row=row, column=16).value
            if cell_value and isinstance(cell_value, str):
                if product == 'HRG' and '1-HRG' in cell_value and '(WD)' not in cell_value:
                    # HRG data starts around row 10, but let's find the actual data start
                    return 10
                elif product == 'WSG' and '2-WSG' in cell_value and '(WD)' not in cell_value:
                    # WSG data starts after HRG data, need to find where HRG ends
                    # For now, let's assume WSG starts around row 20 (this might need adjustment)
                    return 20
        return -1
    
    def find_width_headers_start(self, sheet, header_row=8):
        """Find the starting column for width headers dynamically"""
        # Look for the first column that contains a width value (like "4\"", "6\"", etc.)
        for col in range(1, 50):  # Check columns A through AX
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and isinstance(cell_value, str):
                # Check if it looks like a width header (contains inches symbol)
                if '"' in cell_value and any(char.isdigit() for char in cell_value):
                    return col
        return None  # No fallback - return None if not found
    
    def find_width_headers_end(self, sheet, start_col, header_row=8):
        """Find the ending column for width headers"""
        for col in range(start_col, 50):
            cell_value = sheet.cell(row=header_row, column=col).value
            if not cell_value or not isinstance(cell_value, str) or '"' not in cell_value:
                return col - 1  # Return the previous column (last valid width header)
        return None  # No fallback - return None if not found
    
    def get_price(self, product, finish, size, with_damper=False):
        """Get price for a specific product configuration"""
        try:
            price_type = 'with_damper' if with_damper else 'regular'
            return self.price_data[product][finish][price_type].get(size, 0)
        except KeyError:
            return 0
    
    def get_available_sizes(self, product, finish):
        """Get list of available sizes for a product"""
        try:
            return sorted(list(self.price_data[product][finish]['regular'].keys()))
        except KeyError:
            return []
    
    def find_rounded_size(self, product, finish, width, height):
        """Find the next available size that is >= the given width and height"""
        try:
            available_sizes = self.get_available_sizes(product, finish)
            if not available_sizes:
                return None
            
            # Parse all available sizes to find the best match
            best_match = None
            best_width_diff = float('inf')
            best_height_diff = float('inf')
            
            for size_str in available_sizes:
                # Parse size string like "4" x 6""
                size_match = re.search(r'(\d+(?:\.\d+)?)"?\s*x\s*(\d+(?:\.\d+)?)"?', size_str.lower())
                if size_match:
                    avail_width = float(size_match.group(1))
                    avail_height = float(size_match.group(2))
                    
                    # Check if this size is >= our required size
                    if avail_width >= width and avail_height >= height:
                        width_diff = avail_width - width
                        height_diff = avail_height - height
                        
                        # Find the size with minimum total difference
                        total_diff = width_diff + height_diff
                        if (best_match is None or 
                            total_diff < (best_width_diff + best_height_diff) or
                            (total_diff == (best_width_diff + best_height_diff) and 
                             width_diff < best_width_diff)):
                            best_match = size_str
                            best_width_diff = width_diff
                            best_height_diff = height_diff
            
            return best_match
        except:
            return None

