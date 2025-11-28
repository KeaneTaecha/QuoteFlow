"""
Excel Utilities Module
Contains shared utility functions for reading Excel files, including merged cell handling.
"""


def is_cell_vertically_merged(sheet, row: int, col: int) -> bool:
    """
    Check if a cell is part of a vertically merged range (spans multiple rows).
    
    Args:
        sheet: openpyxl worksheet object
        row: Row number (1-indexed)
        col: Column number (1-indexed)
    
    Returns:
        True if the cell is part of a vertically merged range, False otherwise
    """
    try:
        if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                if hasattr(merged_range, 'min_row'):
                    # It's a CellRange object
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        # Check if it spans multiple rows
                        if merged_range.min_row < merged_range.max_row:
                            return True
                else:
                    # It's a string range like "A1:B2", parse it
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if min_row <= row <= max_row and min_col <= col <= max_col:
                        # Check if it spans multiple rows
                        if min_row < max_row:
                            return True
    except (AttributeError, TypeError, ValueError):
        pass
    return False


def get_cell_value(sheet, row: int, col: int):
    """
    Get cell value, handling merged cells.
    If the cell is part of a merged range, returns the value from the top-left cell.
    
    Args:
        sheet: openpyxl worksheet object
        row: Row number (1-indexed)
        col: Column number (1-indexed)
    
    Returns:
        The cell value, or None if the cell is empty and not part of a merged range
    """
    try:
        # First, try to get the value directly
        cell = sheet.cell(row, col)
        value = cell.value
        
        # If the cell has a value, return it (could be top-left of merged range or regular cell)
        if value is not None:
            return value
        
        # If the cell is None, check if it's part of a merged range
        # In openpyxl, merged cells only have value in the top-left cell
        if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
            try:
                # Iterate through merged ranges
                for merged_range in sheet.merged_cells.ranges:
                    # Check if this cell is within the merged range
                    # merged_range can be a CellRange object or string
                    if hasattr(merged_range, 'min_row'):
                        # It's a CellRange object
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            # Get the value from the top-left cell of the merged range
                            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                            return top_left_cell.value
                    else:
                        # It's a string range like "A1:B2", parse it
                        from openpyxl.utils import range_boundaries
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                        if min_row <= row <= max_row and min_col <= col <= max_col:
                            # Get the value from the top-left cell of the merged range
                            top_left_cell = sheet.cell(min_row, min_col)
                            return top_left_cell.value
            except (AttributeError, TypeError, ValueError) as e:
                # merged_cells might not be accessible or in unexpected format
                pass
        
        # Not part of a merged range, return None
        return None
    except Exception as e:
        # Return None on any error to avoid breaking the process
        return None

