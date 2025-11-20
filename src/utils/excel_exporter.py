"""
Excel Export Module for Quotation System
Generates Excel quotations matching the company template format
"""

import openpyxl
import re
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelQuotationExporter:
    """Handles exporting quotations to Excel format matching company template"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        
    def thai_baht_text(self, amount):
        """Convert amount to Thai baht text"""
        baht = int(amount)
        satang = int(round((amount - baht) * 100))
        
        # Convert baht to Thai text
        baht_text = self._number_to_thai_text(baht)
        
        # Convert satang to Thai text
        satang_text = self._number_to_thai_text(satang)
        
        # Format the result
        if baht == 0 and satang == 0:
            return "จำนวนเงิน ศูนย์บาท ศูนย์สตางค์"
        elif baht == 0:
            return f"จำนวนเงิน ศูนย์บาท {satang_text}สตางค์"
        elif satang == 0:
            return f"จำนวนเงิน {baht_text}บาท"
        else:
            return f"จำนวนเงิน {baht_text}บาท {satang_text}สตางค์"
    
    def _number_to_thai_text(self, number):
        """Convert number to Thai text"""
        if number == 0:
            return "ศูนย์"
        
        # Thai number words
        thai_digits = ['', 'หนึ่ง', 'สอง', 'สาม', 'สี่', 'ห้า', 'หก', 'เจ็ด', 'แปด', 'เก้า']
        thai_units = ['', 'สิบ', 'ร้อย', 'พัน', 'หมื่น', 'แสน', 'ล้าน']
        
        # Special cases for 10-19
        if 10 <= number <= 19:
            if number == 10:
                return "สิบ"
            elif number == 11:
                return "สิบเอ็ด"
            else:
                return "สิบ" + thai_digits[number % 10]
        
        # Handle larger numbers
        result = ""
        position = 0
        
        while number > 0:
            digit = number % 10
            
            if digit != 0:
                if position == 0:  # Units place
                    if digit == 1 and number > 10:  # Special case: เอ็ด for 1 in units when > 10
                        result = "เอ็ด" + result
                    else:
                        result = thai_digits[digit] + result
                elif position == 1:  # Tens place
                    if digit == 1:
                        result = "สิบ" + result
                    elif digit == 2:
                        result = "ยี่สิบ" + result
                    else:
                        result = thai_digits[digit] + "สิบ" + result
                else:  # Hundreds, thousands, etc.
                    if position < len(thai_units):
                        result = thai_digits[digit] + thai_units[position] + result
            
            number //= 10
            position += 1
        
        return result
    
    def get_thai_finishing(self, finish):
        """Convert English finishing name to Thai
        
        Raises:
            ValueError: If "Powder Coated", "Anodized Aluminum", or "Special Color" is in finish but no sub-color is specified (no " - " separator)
        """
        if not finish:
            return ''
        if 'No Finish' in finish:
            # Check if there's a finish_str after "No Finish - "
            if ' - ' in finish:
                finish_str = finish.split(' - ', 1)[1].strip()
                # Return the finish_str if present and not empty, otherwise empty
                return finish_str if finish_str else ''
            else:
                # No finish_str specified, return empty string
                return ''
        elif 'Anodized' in finish:
            # Check if there's a finish_str after "Anodized Aluminum - "
            if ' - ' in finish:
                finish_str = finish.split(' - ', 1)[1].strip()
                # Return the finish_str directly (already in Thai)
                return finish_str
            else:
                raise ValueError(f'Anodized Aluminum finish must include a finish_str. Finish: "{finish}"')
        elif 'Powder Coated' in finish:
            # Check if there's a sub-color after "Powder Coated - "
            if ' - ' in finish:
                color = finish.split(' - ', 1)[1].strip()
                # Return the sub-color directly (already in Thai)
                return color
            else:
                raise ValueError(f'Powder Coated finish must include a sub-color. Finish: "{finish}"')
        elif 'Special Color' in finish:
            # Extract the color name from "Special Color - ColorName" format
            if ' - ' in finish:
                color = finish.split(' - ', 1)[1].strip()
                return color
            else:
                raise ValueError(f'Special Color finish must include a color name. Finish: "{finish}"')
        else:
            return finish  # Return original if no match
    
    def create_excel_quotation(self, quote_data, items, file_path):
        """
        Create Excel quotation using the template file
        
        Args:
            quote_data: Dictionary containing header and footer information
            items: List of quotation items
            file_path: Path to save the Excel file
        """
        # Load the template file
        import sys
        import os
        
        if getattr(sys, 'frozen', False):
            # Running as compiled executable - template is in data folder
            application_path = sys._MEIPASS
            template_path = os.path.join(application_path, 'data', 'quotation_template.xlsx')
        else:
            # Running as script
            template_path = '../data/quotation_template.xlsx'
        
        self.wb = openpyxl.load_workbook(template_path)
        self.ws = self.wb.active
        
        # Store original merged cells info
        self.original_merged_ranges = list(self.ws.merged_cells.ranges)
        
        # Define styles for new content
        normal_font = Font(name='Angsana New', size=14)
        bold_font = Font(name='Angsana New', size=14, bold=True)
        title_font = Font(name='Angsana New', size=14, bold=True, underline='single')
        center_alignment = Alignment(horizontal='center', vertical='bottom')
        left_alignment = Alignment(horizontal='left', vertical='bottom')
        right_alignment = Alignment(horizontal='right', vertical='bottom')
        
        # === STEP 1: CALCULATE EXACT ROWS NEEDED ===
        template_items_start = 14  # Items start at row 14
        template_items_end = 28    # Items area ends at row 28 (footer starts at 29)
        template_available_rows = template_items_end - template_items_start + 1  # 15 rows available
        
        # Calculate exact rows needed (one row per item, no headers)
        total_rows_needed = len(items)
        
        # Calculate how many extra rows to insert
        rows_to_insert = max(0, total_rows_needed - template_available_rows)
        
        # === STEP 2: INSERT ROWS WITH PROPER FORMATTING ===
        if rows_to_insert > 0:
            self._expand_table_for_items(rows_to_insert)
        
        # === POPULATE HEADER SECTION ===
        # Populate the template fields with quotation data
        
        # TO / Company / Tel (template already has labels, just fill the data)
        self._safe_set_cell_value('D5', quote_data.get('to', ''), normal_font)
        self._safe_set_cell_value('D6', quote_data.get('company', ''), normal_font)
        self._safe_set_cell_value('D8', quote_data.get('tel', ''), normal_font)
        self._safe_set_cell_value('H8', quote_data.get('fax', ''), normal_font)
        
        # Quote No / Date / Project (template already has labels)
        self._safe_set_cell_value('N5', quote_data.get('quote_no', ''), normal_font)
        self._safe_set_cell_value('N6', quote_data.get('date', datetime.now().strftime('%Y-%m-%d')), normal_font)
        self._safe_set_cell_value('N7', quote_data.get('project', ''), normal_font)
        
        # === STEP 3: POPULATE PRODUCT INFORMATION ===
        current_row = 14
        item_no = 1
        sub_total = 0
        
        # Get reference row height for consistency (use row 14 as reference)
        reference_row_height = self.ws.row_dimensions[14].height if self.ws.row_dimensions[14].height else None

        for item in items:
            # Set consistent row height for all product rows
            if reference_row_height:
                self.ws.row_dimensions[current_row].height = reference_row_height
            
            if item.get('is_title', False):
                # Title row - no ID number, show title in B column with bold, underline, and left alignment
                self._safe_set_cell_value(f'A{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'B{current_row}', item.get('title', ''), title_font, left_alignment)
                
                # Clear other columns for title (including column D)
                self._safe_set_cell_value(f'D{current_row}', '', normal_font, left_alignment)
                self._safe_set_cell_value(f'G{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'H{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'I{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'J{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'K{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'M{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'O{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'Q{current_row}', '', normal_font, right_alignment)
                # Clear pricing breakdown columns for title rows
                self._safe_set_cell_value(f'V{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'W{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'X{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'Y{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'Z{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'AA{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'AB{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'AC{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'AD{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'AE{current_row}', '', normal_font, right_alignment)
            else:
                # Regular product row
                # NO. in column A (set as integer to avoid "Number Stored as Text" warning)
                self._safe_set_cell_value(f'A{current_row}', item_no, normal_font, center_alignment)
                
                # MODEL in column B (NO MERGING!)
                self._safe_set_cell_value(f'B{current_row}', item.get('product_code', ''), normal_font, left_alignment)
                
                # DETAIL in column D
                detail = item.get('detail', '')
                self._safe_set_cell_value(f'D{current_row}', detail, normal_font, left_alignment)
                
                # FINISHING in column F - Thai text with bottom right alignment
                finish = item.get('finish')
                # Handle None finish - display empty string
                if finish is None:
                    thai_finish = ''
                else:
                    thai_finish = self.get_thai_finishing(finish)
                bottom_right_alignment = Alignment(horizontal='right', vertical='bottom')
                self._safe_set_cell_value(f'F{current_row}', thai_finish, normal_font, bottom_right_alignment)
            
                # Parse size - G(Height) H(x) I(Width) J(Unit) - only for regular items
                size = item.get('size', '')
                rounded_size = item.get('rounded_size', '')
                
                # Check if this is an other_table product (diameter-based)
                # other_table products have rounded_size containing "diameter" or size without "x"
                is_other_table = (
                    'diameter' in str(rounded_size).lower() or 
                    (size and 'x' not in size and size.strip())
                )
                
                def parse_dimension(value_str, field_name):
                    """Parse dimension value and detect unit. Returns (numeric_value, unit_string)."""
                    if not value_str:
                        raise ValueError(f"Empty {field_name} value")
                    value_str = value_str.strip()
                    
                    # Detect unit (check in order: longer units first to avoid partial matches)
                    value_lower = value_str.lower()
                    has_cm = 'cm' in value_lower and 'mm' not in value_lower  # cm but not part of mm
                    has_mm = 'mm' in value_lower
                    has_m = re.search(r'\b(m|meter|metre)\b', value_lower) and 'mm' not in value_lower and 'cm' not in value_lower
                    has_ft = re.search(r'\b(ft|feet|foot)\b', value_lower) or "'" in value_str
                    has_quote = '"' in value_str
                    
                    # Extract numeric value (remove all unit symbols)
                    cleaned = value_str.replace('"', '').replace("'", '').replace('mm', '').replace('MM', '').replace('Mm', '').replace('cm', '').replace('CM', '').replace('Cm', '').replace('m', '').replace('M', '').replace('ft', '').replace('FT', '').replace('Ft', '').replace('feet', '').replace('Feet', '').replace('foot', '').replace('Foot', '').replace('meter', '').replace('Meter', '').replace('metre', '').replace('Metre', '').strip()
                    match = re.search(r'-?\d+\.?\d*', cleaned)
                    if not match:
                        raise ValueError(f"Could not extract numeric value from {field_name} '{value_str}'")
                    
                    try:
                        numeric_value = float(match.group())
                    except ValueError as e:
                        raise ValueError(f"Could not convert {field_name} '{value_str}' to float: {e}")
                    
                    # Determine unit (check in order of specificity)
                    if has_cm:
                        return numeric_value, 'cm'
                    elif has_mm:
                        return numeric_value, 'mm'
                    elif has_m:
                        return numeric_value, 'm'
                    elif has_ft:
                        return numeric_value, 'ft'
                    elif has_quote:
                        return numeric_value, '"'
                    else:
                        # No explicit unit - default to mm if value > 100, otherwise inches
                        return numeric_value, 'mm' if numeric_value > 100 else '"'
                
                # Handle other_table products (diameter-based) differently
                if is_other_table:
                    # For other_table products, size is just the diameter (e.g., "6.0" or "6.0"")
                    # Parse it as a single dimension
                    diameter_val, diameter_unit = parse_dimension(size, 'diameter')
                    
                    # Helper to format numbers (remove .0 for whole numbers)
                    def format_number(num):
                        return int(num) if num == int(num) else num
                    
                    # Display diameter in height column (G), leave width column (I) empty or show "x"
                    # Format: diameter in G, "x" in H, empty in I, unit in J
                    if diameter_unit == '"':
                        diameter_display = format_number(diameter_val)
                        unit_display = 'in'
                    elif diameter_unit == 'cm':
                        diameter_display = format_number(diameter_val)
                        unit_display = 'cm'
                    elif diameter_unit == 'm':
                        diameter_display = format_number(diameter_val)
                        unit_display = 'm'
                    elif diameter_unit == 'ft':
                        diameter_display = format_number(diameter_val)
                        unit_display = 'ft'
                    else:
                        diameter_display = format_number(diameter_val)
                        unit_display = 'mm'
                    
                    # For other_table, we might want to show diameter in a specific way
                    # Based on the template, we'll put diameter in height column (G)
                    self._safe_set_cell_value(f'G{current_row}', diameter_display, normal_font, center_alignment)
                    self._safe_set_cell_value(f'H{current_row}', '', normal_font, center_alignment)  # Empty for diameter
                    self._safe_set_cell_value(f'I{current_row}', '', normal_font, center_alignment)  # Empty for diameter
                    self._safe_set_cell_value(f'J{current_row}', unit_display, normal_font, center_alignment)
                else:
                    # Regular products with WIDTH x HEIGHT format
                    parts = size.split('x')
                    if len(parts) != 2:
                        raise ValueError(f"Invalid size format (expected 'WIDTH x HEIGHT'): '{size}'")
                    
                    width_part = parts[0].strip()
                    height_part = parts[1].strip()
                    
                    # Check if width contains "Slot" - preserve it in display
                    has_slot = 'Slot' in width_part or 'slot' in width_part
                    
                    width_val, width_unit = parse_dimension(width_part, 'width')
                    height_val, height_unit = parse_dimension(height_part, 'height')
                    
                    # Helper to format numbers (remove .0 for whole numbers)
                    def format_number(num):
                        return int(num) if num == int(num) else num
                    
                    # Helper to get unit display symbol
                    def get_unit_display(unit):
                        """Convert unit string to display symbol."""
                        if unit == '"':
                            return 'in'
                        elif unit == 'mm':
                            return 'mm'
                        elif unit == 'cm':
                            return 'cm'
                        elif unit == 'm':
                            return 'm'
                        elif unit == 'ft':
                            return 'ft'
                        else:
                            return 'in'  # Default
                    
                    # Helper to format dimension with unit if needed
                    def format_dimension_with_unit(value, unit):
                        """Format dimension value, adding unit symbol if unit is not standard."""
                        num = format_number(value)
                        if unit in ['mm', 'cm', 'm', 'ft']:
                            return f'{num}{unit}'
                        elif unit == '"':
                            return f'{num}"'
                        else:
                            return num
                    
                    # Display logic: if same unit, show numbers only; if different, show units after numbers
                    # Special handling for "Slot" in width
                    if has_slot:
                        # Slot has no unit, always display as "7Slot" format
                        width_display = f'{format_number(width_val)}Slot'
                        # Use height unit for unit_display since Slot has no unit
                        height_display = format_number(height_val)
                        unit_display = get_unit_display(height_unit)
                    elif width_unit == height_unit:
                        # Same unit: show numbers only, put unit in column J
                        width_display = format_number(width_val)
                        height_display = format_number(height_val)
                        unit_display = get_unit_display(width_unit)
                    else:
                        # Different units: show units after numbers, leave column J empty
                        width_display = format_dimension_with_unit(width_val, width_unit)
                        height_display = format_dimension_with_unit(height_val, height_unit)
                        
                        unit_display = ''  # Empty when units are different (units shown with numbers)
                    
                    self._safe_set_cell_value(f'G{current_row}', width_display, normal_font, center_alignment)
                    self._safe_set_cell_value(f'H{current_row}', 'x', normal_font, center_alignment)
                    self._safe_set_cell_value(f'I{current_row}', height_display, normal_font, center_alignment)
                    self._safe_set_cell_value(f'J{current_row}', unit_display, normal_font, center_alignment)
                
                # QTY in column K
                quantity = int(item.get('quantity', 1))
                self._safe_set_cell_value(f'K{current_row}', quantity, normal_font, center_alignment)
                
                # Discount in column O (unchanged)
                discount = item.get('discount', 0)
                if discount > 0:
                    # Set as decimal number (0.1 for 10%) and apply percentage format
                    cell = self.ws[f'O{current_row}']
                    cell.value = discount
                    cell.font = normal_font
                    cell.alignment = center_alignment
                    cell.number_format = '0%'  # Format as percentage
                else:
                    self._safe_set_cell_value(f'O{current_row}', '', normal_font, center_alignment)
                
                # UNIT PRICE in column M - links to AE column (รวมทั้งหมด)
                cell_m = self.ws[f'M{current_row}']
                cell_m.value = f'=AE{current_row}'
                cell_m.font = normal_font
                cell_m.alignment = right_alignment
                cell_m.number_format = '0.00'  # Format to 2 decimal places
                
                # AMOUNT in column Q - ราคาต่อหน่วย x จำนวน (M * K)
                cell_q = self.ws[f'Q{current_row}']
                cell_q.value = f'=M{current_row}*K{current_row}'
                cell_q.font = normal_font
                cell_q.alignment = right_alignment
                cell_q.number_format = '0.00'  # Format to 2 decimal places
                
                # Calculate item_total for footer (using AE value, but we'll calculate it after AE is set)
                # We'll need to recalculate this after AE column is populated
                unit_price = float(item.get('unit_price', 0))
                if discount > 0:
                    discounted_price = unit_price * (1 - discount)
                else:
                    discounted_price = unit_price
                item_total = discounted_price * quantity
                sub_total += item_total
                
                # === PRICING BREAKDOWN TABLE STARTING AT COLUMN V ===
                # Column V: List (table price)
                table_price = item.get('table_price', 0.0)
                cell_v = self.ws[f'V{current_row}']
                cell_v.value = table_price
                cell_v.font = normal_font
                cell_v.alignment = right_alignment
                cell_v.number_format = '0.00'  # Format to 2 decimal places
                
                # Column W: พ่นส๊ (List * finish multiplier = price_after_finish)
                price_after_finish = item.get('price_after_finish', 0.0)
                ins_price = item.get('ins_price', 0.0)
                filter_price = item.get('filter_price', 0.0)
                
                # Use formula: =V{row} * (price_after_finish/table_price) if table_price > 0
                # If table_price is 0, use IF formula to avoid division by zero
                cell_w = self.ws[f'W{current_row}']
                if table_price > 0:
                    finish_multiplier = price_after_finish / table_price
                    cell_w.value = f'=V{current_row}*{finish_multiplier}'
                else:
                    # If table_price is 0, just use price_after_finish directly
                    cell_w.value = f'=IF(V{current_row}=0,{price_after_finish},V{current_row}*({price_after_finish}/V{current_row}))'
                cell_w.font = normal_font
                cell_w.alignment = right_alignment
                cell_w.number_format = '0.00'  # Format to 2 decimal places
                
                # Column X: INS price
                cell_x = self.ws[f'X{current_row}']
                cell_x.value = ins_price
                cell_x.font = normal_font
                cell_x.alignment = right_alignment
                cell_x.number_format = '0.00'  # Format to 2 decimal places
                
                # Column Y: Filter price
                cell_y = self.ws[f'Y{current_row}']
                cell_y.value = filter_price
                cell_y.font = normal_font
                cell_y.alignment = right_alignment
                cell_y.number_format = '0.00'  # Format to 2 decimal places
                
                # Column Z: subtotal (พ่นส๊ + INS + Filter)
                cell_z = self.ws[f'Z{current_row}']
                cell_z.value = f'=W{current_row}+X{current_row}+Y{current_row}'
                cell_z.font = normal_font
                cell_z.alignment = right_alignment
                cell_z.number_format = '0.00'  # Format to 2 decimal places
                
                # Column AA: Discount (from column O)
                # Copy discount value from column O to column AA for pricing breakdown
                cell_aa = self.ws[f'AA{current_row}']
                if discount > 0:
                    cell_aa.value = discount
                    cell_aa.font = normal_font
                    cell_aa.alignment = right_alignment
                    cell_aa.number_format = '0%'  # Format as percentage
                else:
                    cell_aa.value = 0
                    cell_aa.font = normal_font
                    cell_aa.alignment = right_alignment
                    cell_aa.number_format = '0%'
                
                # Column AB: Total (subtotal * (1 - Discount) from AA column)
                # When discount is 0%, Total = Subtotal; when discount > 0%, Total = Subtotal * (1 - Discount)
                cell_ab = self.ws[f'AB{current_row}']
                cell_ab.value = f'=Z{current_row}*(1-AA{current_row})'
                cell_ab.font = normal_font
                cell_ab.alignment = right_alignment
                cell_ab.number_format = '0.00'  # Format to 2 decimal places
                
                # Column AC: ค่าสี (Color cost) - left blank for manual entry
                cell_ac = self.ws[f'AC{current_row}']
                cell_ac.value = ''
                cell_ac.font = normal_font
                cell_ac.alignment = right_alignment
                cell_ac.number_format = '0.00'  # Format to 2 decimal places
                
                # Column AD: ค่าขนส่ง (Shipping cost) - left blank for manual entry
                cell_ad = self.ws[f'AD{current_row}']
                cell_ad.value = ''
                cell_ad.font = normal_font
                cell_ad.alignment = right_alignment
                cell_ad.number_format = '0.00'  # Format to 2 decimal places
                
                # Column AE: รวมทั้งหมด (Grand Total) = Total + ค่าสี + ค่าขนส่ง
                cell_ae = self.ws[f'AE{current_row}']
                cell_ae.value = f'=AB{current_row}+AC{current_row}+AD{current_row}'
                cell_ae.font = normal_font
                cell_ae.alignment = right_alignment
                cell_ae.number_format = '0.00'  # Format to 2 decimal places
                
                item_no += 1
            
            current_row += 1
        
        # === POPULATE FOOTER SECTION ===
        # Calculate totals and populate template footer fields
        # Footer starts at row 29 in template, but shifted down by inserted rows
        footer_start_row = 29 + rows_to_insert
        
        # Calculate the last item row (current_row - 1 after the loop)
        last_item_row = current_row - 1
        first_item_row = 14
        
        # Sub total - use SUM formula for column Q
        cell_subtotal = self.ws[f'Q{footer_start_row}']
        cell_subtotal.value = f'=SUM(Q{first_item_row}:Q{last_item_row})'
        cell_subtotal.font = bold_font
        cell_subtotal.alignment = right_alignment
        
        # VAT - 7% of subtotal
        cell_vat = self.ws[f'Q{footer_start_row + 1}']
        cell_vat.value = f'=Q{footer_start_row}*0.07'
        cell_vat.font = bold_font
        cell_vat.alignment = right_alignment
        
        # Grand total - subtotal + VAT
        grand_total = sub_total + (sub_total * 0.07)  # For Thai baht text calculation
        self._safe_set_cell_value(f'A{footer_start_row + 2}', self.thai_baht_text(grand_total), normal_font)
        cell_grand_total = self.ws[f'Q{footer_start_row + 2}']
        cell_grand_total.value = f'=Q{footer_start_row}+Q{footer_start_row + 1}'
        cell_grand_total.font = Font(name='Angsana New', size=14, bold=True)
        cell_grand_total.alignment = right_alignment
        
        # Footer information - populate template fields
        footer_row = footer_start_row + 4
        self._safe_set_cell_value(f'D{footer_row}', quote_data.get('remarks', '1. ยืนราคา 60 วัน'), normal_font)
        self._safe_set_cell_value(f'M{footer_row}', quote_data.get('quoted_by_name', ''), normal_font, center_alignment)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('payment_term', 'เครดิต 30 วัน'), normal_font)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('delivery_place', ''), normal_font)
        self._safe_set_cell_value(f'O{footer_row}', quote_data.get('purchased_by', ''), normal_font, center_alignment)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('delivery_date', ''), normal_font)
        
        # Save the workbook
        self.wb.save(file_path)
        return True
    
    def _expand_table_for_items(self, rows_to_insert):
        """
        Expand the table to accommodate more items by inserting rows with proper formatting
        
        Args:
            rows_to_insert: Number of rows to insert
        """
        if rows_to_insert <= 0:
            return
        
        # Insert rows INSIDE the table, just before the footer section
        insert_position = 28  # Insert before the footer section
        
        # === STEP 1: Handle merged cells in footer section ===
        # Store merged cell ranges that are at or below the insert position
        footer_merged_cells = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # Check if this merged range starts at or after the insert position
            if merged_range.min_row >= insert_position:
                # Store the range information
                footer_merged_cells.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
                # Unmerge this range
                self.ws.unmerge_cells(str(merged_range))
        
        # === STEP 2: Insert the rows ===
        reference_row = 27  # Reference row to copy formatting from
        self.ws.insert_rows(insert_position, rows_to_insert)
        
        # === STEP 3: Copy formatting to newly inserted rows ===
        # Get reference row height for consistency
        reference_row_height = self.ws.row_dimensions[reference_row].height
        
        for i in range(rows_to_insert):
            target_row = insert_position + i
            
            # Set consistent row height
            if reference_row_height:
                self.ws.row_dimensions[target_row].height = reference_row_height
            
            # Copy formatting for each column
            for col in range(1, 19):  # A to Q (columns 1-17, extended range)
                source_cell = self.ws.cell(row=reference_row, column=col)
                target_cell = self.ws.cell(row=target_row, column=col)
                
                # Copy font
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color
                    )
                
                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )
                
                # Copy border
                if source_cell.border:
                    target_cell.border = Border(
                        left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
                        right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
                        top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
                        bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                    )
                
                # Copy fill
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
                
                # Copy number format (important for unit prices)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
        
        # === STEP 4: Re-merge cells in footer at new positions ===
        for merged_info in footer_merged_cells:
            # Calculate new row positions (shifted down by rows_to_insert)
            new_min_row = merged_info['min_row'] + rows_to_insert
            new_max_row = merged_info['max_row'] + rows_to_insert
            
            # Reconstruct the range string
            start_cell = f"{get_column_letter(merged_info['min_col'])}{new_min_row}"
            end_cell = f"{get_column_letter(merged_info['max_col'])}{new_max_row}"
            range_string = f"{start_cell}:{end_cell}"
            
            # Re-merge the cells at the new position
            try:
                self.ws.merge_cells(range_string)
            except Exception as e:
                print(f"Warning: Could not re-merge cells {range_string}: {e}")

    def _safe_set_cell_value(self, cell_ref, value, font=None, alignment=None):
        """Safely set cell value, handling merged cells"""
        try:
            cell = self.ws[cell_ref]
            
            # Check if this cell is part of a merged range
            if self._is_merged_cell(cell):
                # Find the top-left cell of the merged range
                for merged_range in self.ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        # Get the top-left cell of the merged range
                        top_left = merged_range.min_row, merged_range.min_col
                        top_left_cell = self.ws.cell(row=top_left[0], column=top_left[1])
                        top_left_cell.value = value
                        if font:
                            top_left_cell.font = font
                        if alignment:
                            top_left_cell.alignment = alignment
                        break
            else:
                # Regular cell, set value directly
                cell.value = value
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                    
        except Exception as e:
            print(f"Warning: Could not set value for {cell_ref}: {e}")
    
    def _safe_merge_cells(self, range_string):
        """Safely merge cells, checking for conflicts"""
        try:
            # Check if any cell in the range is already merged
            start_col, start_row, end_col, end_row = openpyxl.utils.range_boundaries(range_string)
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell_ref = openpyxl.utils.get_column_letter(col) + str(row)
                    for merged_range in self.ws.merged_cells.ranges:
                        if cell_ref in merged_range:
                            # This cell is already merged, skip merging
                            return False
            
            self.ws.merge_cells(range_string)
            return True
        except Exception as e:
            print(f"Warning: Could not merge cells {range_string}: {e}")
            return False
    
    def _is_merged_cell(self, cell):
        """Check if a cell is part of a merged range"""
        try:
            cell_ref = cell.coordinate
            for merged_range in self.ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    return True
            return False
        except:
            return False
    