"""
Excel Importer Module
Handles importing items from Excel files into quotations.
"""

import re
import openpyxl
from typing import List, Dict, Optional, Tuple
from utils.price_calculator import PriceCalculator
from utils.filter_utils import get_filter_price
from utils.product_utils import (
    extract_product_flags_and_filter,
    validate_product_exists,
    extract_slot_number_from_model, get_product_type_flags
)
from utils.quote_utils import build_quote_item
from utils.product_utils import parse_dimension_with_unit


class ExcelItemImporter:
    """Handles importing items from Excel files"""
    
    def __init__(self, price_calculator: PriceCalculator, available_models: List[str]):
        """
        Initialize the Excel importer
        
        Args:
            price_calculator: PriceCalculator instance for database access
            available_models: List of available product models
        """
        self.price_calculator = price_calculator
        self.available_models = available_models
    
    def parse_excel_file(self, file_path: str, progress_callback=None) -> List[Dict]:
        """
        Parse Excel file and extract items based on Model, Detail, Width, Height, Unit, Quantity, Finish columns
        
        Args:
            file_path: Path to the Excel file
            progress_callback: Optional callback function(progress_percent_0_to_100, status_text) to report progress
                               The callback receives internal progress 0-100%, caller should map to overall progress
        """
        if progress_callback:
            progress_callback(5, 'Loading Excel file...')
        
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active  # Use the first/active sheet
            
            if progress_callback:
                progress_callback(10, 'Searching for header row...')
            
            # Find header row by searching for keywords
            header_row = None
            column_mapping = {}
            
            # Search for header row (search first 20 rows)
            for row in range(1, min(21, sheet.max_row + 1)):
                # Check all columns in this row for header keywords
                row_mapping = {}
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row, col).value
                    if cell_value is None:
                        continue
                    
                    cell_str = str(cell_value).strip().lower()
                    
                    # Check for keywords
                    if 'model' in cell_str and 'model' not in row_mapping:
                        row_mapping['model'] = col
                    elif 'detail' in cell_str and 'detail' not in row_mapping:
                        row_mapping['detail'] = col
                    elif 'width' in cell_str and 'width' not in row_mapping:
                        row_mapping['width'] = col
                    elif 'height' in cell_str and 'height' not in row_mapping:
                        row_mapping['height'] = col
                    elif 'unit' in cell_str and 'unit' not in row_mapping:
                        row_mapping['unit'] = col
                    elif 'quantity' in cell_str and 'quantity' not in row_mapping:
                        row_mapping['quantity'] = col
                    elif 'finish' in cell_str and 'finish' not in row_mapping:
                        row_mapping['finish'] = col
                    elif 'discount' in cell_str and 'discount' not in row_mapping:
                        row_mapping['discount'] = col
                
                # If we found Model column and at least a few other columns, use this row
                if 'model' in row_mapping and len(row_mapping) >= 2:
                    header_row = row
                    column_mapping = row_mapping
                    break
            
            if header_row is None or 'model' not in column_mapping:
                raise ValueError("Could not find required 'Model' column in Excel file")
            
            # Extract items from rows below header
            items = []
            model_col = column_mapping.get('model')
            detail_col = column_mapping.get('detail')
            width_col = column_mapping.get('width')
            height_col = column_mapping.get('height')
            unit_col = column_mapping.get('unit')
            quantity_col = column_mapping.get('quantity')
            finish_col = column_mapping.get('finish')
            discount_col = column_mapping.get('discount')
            
            # Calculate total rows to process
            total_rows = sheet.max_row - header_row
            processed_rows = 0
            
            # Process rows below header
            for row in range(header_row + 1, sheet.max_row + 1):
                model_value = self._get_cell_value(sheet, row, model_col)
                
                # Get other column values
                detail_value = self._get_cell_value(sheet, row, detail_col) if detail_col else None
                width_value = self._get_cell_value(sheet, row, width_col) if width_col else None
                height_value = self._get_cell_value(sheet, row, height_col) if height_col else None
                unit_value = self._get_cell_value(sheet, row, unit_col) if unit_col else None
                quantity_value = self._get_cell_value(sheet, row, quantity_col) if quantity_col else None
                finish_value = self._get_cell_value(sheet, row, finish_col) if finish_col else None
                discount_value = self._get_cell_value(sheet, row, discount_col) if discount_col else None
                
                # Check if model is empty (blank row) or if this is a title (Model has text but other columns are empty)
                model_str = str(model_value).strip() if model_value is not None else ''
                is_blank_row = model_str == ''
                
                has_detail = detail_value is not None and str(detail_value).strip() != ''
                has_width = width_value is not None and str(width_value).strip() != ''
                has_height = height_value is not None and str(height_value).strip() != ''
                has_quantity = quantity_value is not None and str(quantity_value).strip() != ''
                has_finish = finish_value is not None and str(finish_value).strip() != ''
                
                if is_blank_row or (not has_detail and not has_width and not has_height and not has_quantity and not has_finish):
                    # This is a title (blank row or model has text but other columns are empty)
                    items.append({
                        'is_title': True,
                        'title': model_str,
                        'product_code': '',
                        'size': '',
                        'finish': '',
                        'quantity': 0,
                        'unit_price': 0,
                        'discount': 0,
                        'discounted_unit_price': 0,
                        'total': 0,
                        'rounded_size': None,
                        'detail': ''
                    })
                else:
                    # This is a product item
                    # Handle finish: if empty or whitespace, set to None to ensure multiplier is 1.0
                    finish_str = str(finish_value).strip() if finish_value else ''
                    finish_for_item = None if not finish_str else finish_str
                    
                    item = {
                        'model': model_str,
                        'detail': str(detail_value).strip() if detail_value else '',
                        'width': width_value,
                        'height': height_value,
                        'unit': str(unit_value).strip().lower() if unit_value else 'inches',
                        'quantity': self._parse_number(quantity_value) if quantity_value else 1,
                        'finish': finish_for_item,
                        'discount': self._parse_discount(discount_value) if discount_value else 0.0
                    }
                    items.append(item)
                
                processed_rows += 1
                # Update progress every 10 rows or at the end
                if progress_callback and total_rows > 0 and (processed_rows % 10 == 0 or processed_rows == total_rows):
                    progress = 15 + int((processed_rows / total_rows) * 85)  # 15% to 100% internal
                    progress = min(progress, 100)
                    progress_callback(progress, f'Reading row {row} of {sheet.max_row}...')
            
            if progress_callback:
                progress_callback(100, 'Parsing complete!')
            
            return items
        finally:
            # Always close the workbook, even if an exception occurred
            if wb is not None:
                wb.close()
    
    def _get_cell_value(self, sheet, row, col):
        """Get cell value safely, returning None if cell doesn't exist"""
        if col is None:
            return None
        try:
            return sheet.cell(row, col).value
        except:
            return None
    
    def _parse_number(self, value):
        """Parse a number from a cell value"""
        if value is None:
            return 1
        try:
            # Try to convert to float first, then int
            num = float(str(value).strip())
            return int(num) if num.is_integer() else num
        except:
            return 1
    
    def _parse_discount(self, value):
        """Parse discount value, handling both percentage strings (e.g., "90%") and decimal values (e.g., 0.9)
        
        If value is a decimal < 1 (like 0.9), it's treated as Excel's percentage format and multiplied by 100.
        If value is >= 1, it's used as-is (already a percentage).
        If value contains "%", the symbol is removed and the number is used.
        """
        if value is None:
            return 0.0
        try:
            # If it's already a number, check if it's a decimal percentage
            if isinstance(value, (int, float)):
                # If value is < 1, it's likely Excel's decimal format (0.9 = 90%)
                if 0 < value < 1:
                    return value * 100
                # Otherwise use as-is (already a percentage value like 90)
                return float(value)
            
            # If it's a string, remove % and parse
            value_str = str(value).replace('%', '').strip()
            num = float(value_str)
            return num
        except:
            return 0.0
    
    def _create_invalid_item(self, item_data, error_message):
        """Create an invalid item entry for display in the quote"""
        model = item_data.get('model', 'Unknown').strip()
        detail = item_data.get('detail', '').strip()
        width_value = item_data.get('width')
        height_value = item_data.get('height')
        quantity = item_data.get('quantity', 1)
        finish_value = item_data.get('finish')
        
        # Create size string from available dimensions
        size_parts = []
        if width_value is not None:
            size_parts.append(f"W: {width_value}")
        if height_value is not None:
            size_parts.append(f"H: {height_value}")
        size_str = " x ".join(size_parts) if size_parts else "N/A"
        
        # Create finish string
        finish_str = 'INVALID'
        if finish_value:
            finish_str = f"INVALID ({finish_value})"
        
        return {
            'is_invalid': True,
            'product_code': model,
            'size': size_str,
            'finish': finish_str,
            'quantity': quantity,
            'unit_price': 0,
            'discount': 0,
            'discounted_unit_price': 0,
            'total': 0,
            'rounded_size': None,
            'detail': detail,
            'error_message': error_message
        }
    
    def add_item_from_excel(self, item_data):
        """Add an item from Excel data to the quote. Returns dict with 'success' and 'error' keys."""
        # Validate database is initialized first
        if not self.available_models:
            return {'success': False, 'error': 'Price database not initialized'}
        
        model = item_data.get('model', '').strip()
        detail = item_data.get('detail', '').strip()
        width_value = item_data.get('width')
        height_value = item_data.get('height')
        unit_str = item_data.get('unit', 'inches').lower()
        quantity = item_data.get('quantity', 1)
        finish_from_excel = item_data.get('finish')  # Finish from Excel file
        discount = item_data.get('discount', 0.0)  # Discount percentage (0-100), default 0
        
        # Determine default unit - validate unit is compatible
        unit_str_lower = str(unit_str).lower().strip()
        if 'mm' in unit_str_lower or 'millimeter' in unit_str_lower:
            default_unit = 'millimeters'
        elif 'cm' in unit_str_lower or 'centimeter' in unit_str_lower:
            default_unit = 'centimeters'
        elif unit_str_lower == 'm' or 'meter' in unit_str_lower:
            default_unit = 'meters'
        elif unit_str_lower == 'ft' or unit_str_lower == "'" or 'foot' in unit_str_lower or 'feet' in unit_str_lower:
            default_unit = 'feet'
        elif 'inch' in unit_str_lower or unit_str_lower == '"' or unit_str_lower == 'in' or unit_str_lower == '':
            # Empty string defaults to inches (for backward compatibility when unit column is empty)
            default_unit = 'inches'
        else:
            # Unit is provided but not compatible - raise error
            return {'success': False, 'error': f'Incompatible unit "{unit_str}". Supported units are: inches (or "), millimeters (or mm), centimeters (or cm), meters (or m), feet (or ft)'}
        
        # Extract filter suffix, WD, and INS if present
        # Use combined extraction to handle WD, INS, and filter
        base_model, has_wd_from_name, has_ins_from_name, filter_type = extract_product_flags_and_filter(model)
        
        # Validate and get normalized product info (with filter validation)
        # Pass pre-extracted values to avoid redundant extraction
        product_exists, product, has_wd_from_db, error_msg = validate_product_exists(
            base_model, self.available_models, self.price_calculator, filter_type,
            has_wd=has_wd_from_name
        )
        
        # Use WD from name if specified, otherwise use from database validation
        has_wd = has_wd_from_name or has_wd_from_db
        
        if not product_exists:
            return {'success': False, 'error': error_msg or f'Product "{model}" not found in database'}
        
        # Get product type flags using consolidated helper
        has_no_dimensions, has_price_per_foot, has_price_per_sq_in, is_other_table = get_product_type_flags(self.price_calculator, product)
        
        # Parse width and height with unit handling
        width, width_unit = parse_dimension_with_unit(width_value)
        height, height_unit = parse_dimension_with_unit(height_value)
        
        # If no unit detected, use default_unit
        if width_unit is None:
            width_unit = default_unit
        if height_unit is None:
            height_unit = default_unit
        
        # For price_per_foot or price_per_sq_in products, both width and height are required (unless has_no_dimensions)
        if (has_price_per_foot or has_price_per_sq_in) and not has_no_dimensions:
            if width is None or height is None:
                return {'success': False, 'error': f'Width and height are required for price_per_foot/price_per_sq_in product {product}. Please provide both dimensions in the Excel file.'}
        
        # Allow products with no dimensions to be imported without width/height
        if width is None and height is None and not has_no_dimensions:
            # Check if this is an "other table" product (diameter-based)
            # For now, we'll skip items without dimensions unless they have no_dimensions flag
            return {'success': False, 'error': 'Missing dimensions (width and height)'}
        
        # Get available finishes
        finishes = self.price_calculator.get_available_finishes(product)
        if not finishes:
            return {'success': False, 'error': f'No finishes available for product "{product}"'}
        
        # Determine which finish to use and extract special color multiplier
        finish, special_color_multiplier = self._match_finish(finish_from_excel, finishes)
        
        # Check if finish column was empty - if so, use "No Finish" if available
        finish_was_empty = finish_from_excel is None or (isinstance(finish_from_excel, str) and not finish_from_excel.strip())
        if finish is None and finish_was_empty and 'No Finish' in finishes:
            finish = 'No Finish'
            special_color_multiplier = None
        
        # If finish was provided but doesn't match, return error
        if finish is None:
            return {'success': False, 'error': f'Finish "{finish_from_excel}" not available for product "{product}". Available finishes: {", ".join(finishes)}'}
        
        # Extract slot number for no-dimension products
        slot_number = None
        if has_no_dimensions:
            slot_number = extract_slot_number_from_model(model)
        
        # Use shared function to build quote item
        quote_item, error = build_quote_item(
            price_calculator=self.price_calculator,
            product=product,
            finish=finish,
            quantity=quantity,
            has_wd=has_wd,
            has_price_per_foot=has_price_per_foot,
            has_price_per_sq_in=has_price_per_sq_in,
            is_other_table=is_other_table,
            width=width,
            height=height,
            width_unit=width_unit,
            height_unit=height_unit,
            filter_type=filter_type,
            discount=discount,
            special_color_multiplier=special_color_multiplier,
            detail=detail,
            has_ins=has_ins_from_name,
            has_no_dimensions=has_no_dimensions,
            slot_number=slot_number
        )
        
        if error:
            return {'success': False, 'error': error}
        
        return {'success': True, 'item': quote_item, 'error': None}
    
    def _match_finish(self, finish_from_excel, finishes):
        """Match finish from Excel with available finishes and extract multiplier for special colors
        
        Returns:
            tuple: (finish_name, multiplier) where multiplier is 1.0 for non-special colors
        """
        # If finish is None, empty string, or whitespace, return None (will be handled by caller)
        if not finish_from_excel or (isinstance(finish_from_excel, str) and not finish_from_excel.strip()):
            return (None, 1.0)
        
        finish_str = str(finish_from_excel).strip()
        # If after stripping it's empty, return None (will be handled by caller)
        if not finish_str:
            return (None, 1.0)
        
        finish_lower = finish_str.lower()  
        # Define keywords for Powder Coated finish (used for both substring matching and exact sub-color matching)
        powder_keywords = ['ขาวนวล', 'ขาวด้าน', 'ขาวฟ้า', 'ขาวควันบุหรี่', 'ดำด้าน', 'ดำเงา', 'บรอนซ์', 'สีดำ', 'พ่นดำ', 'สีอบขาว', 'สีพ่นขาว']
        anodized_keywords = ['anodized', 'aluminum', 'anodized aluminum', 'anodised', 'aluminium', 'anodised aluminium', 'สีอลูมิเนียม']
        
        # Check for "No Finish" keywords first
        no_finish_keywords = ['no finish', 'nofinish', 'no_finish', 'raw', 'unfinished', 'สังกะสี', 'stainless steel', 'ไม่ทำสี']

        if any(keyword in finish_lower for keyword in no_finish_keywords):
            if 'No Finish' in finishes:
                return (f"No Finish - {finish_str}", None)
            else:
                return (None, 1.0)
     
        # Check for Powder Coated
        elif any(keyword in finish_lower for keyword in powder_keywords):
            if 'Powder Coated' in finishes:
                return (f"Powder Coated - {finish_str}", None)
            else:
                # Powder Coated keywords matched but finish not available
                return (None, 1.0)
        
        # Check for Anodized Aluminum
        elif any(keyword in finish_lower for keyword in anodized_keywords):
            if 'Anodized Aluminum' in finishes:
                return (f"Anodized Aluminum - {finish_str}", None)
            else:
                # Anodized Aluminum keywords matched but finish not available
                return (None, 1.0)
        
        # Default: treat as Special Color and extract multiplier
        else:
            if 'Special Color' or 'สีพิเศษ' in finishes:
                # Extract multiplier if present (format: "Color, Multiplier")
                multiplier = None
                color_name = finish_str
                if ',' in finish_str:
                    parts = finish_str.split(',', 1)
                    if len(parts) == 2:
                        color_name = parts[0].strip()
                        try:
                            multiplier = float(parts[1].strip())
                        except ValueError:
                            pass  # Keep default 1.0 if invalid
                
                # Return special color with extracted name and multiplier
                return (f"Special Color - {color_name}", multiplier)
            else:
                # Special Color not available in finishes
                return (None, 1.0)
    

