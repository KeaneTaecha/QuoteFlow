"""
Excel Export Module for Quotation System
Generates Excel quotations matching the company template format
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelQuotationExporter:
    """Handles exporting quotations to Excel format matching company template"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        
    def thai_baht_text(self, amount):
        """Convert amount to Thai baht text"""
        # This is a simplified version - you may want to use a proper Thai number library
        # For now, we'll use a basic conversion
        baht = int(amount)
        satang = int((amount - baht) * 100)
        
        # Thai number words (simplified - you may want to use pythainlp library)
        thai_numbers = {
            0: 'ศูนย์', 1: 'หนึ่ง', 2: 'สอง', 3: 'สาม', 4: 'สี่',
            5: 'ห้า', 6: 'หก', 7: 'เจ็ด', 8: 'แปด', 9: 'เก้า',
            10: 'สิบ', 20: 'ยี่สิบ', 100: 'ร้อย', 1000: 'พัน',
            10000: 'หมื่น', 100000: 'แสน', 1000000: 'ล้าน'
        }
        
        # For demonstration, return a placeholder
        # In production, use pythainlp or similar library
        return f"จำนวนเงิน {baht:,} บาท {satang:02d} สตางค์"
    
    def get_thai_finishing(self, finish):
        """Convert English finishing name to Thai"""
        if 'Anodized' in finish:
            return 'สีอลูมิเนียม'
        elif 'White' in finish:
            return 'สีขาว'
        elif 'Other Paint' in finish:
            return 'สีอื่นๆ'
        else:
            return finish  # Return original if no match
    
    def create_excel_quotation(self, quote_data, items, file_path):
        """
        Create Excel quotation using the template file
        
        Args:
            quote_data: Dictionary containing header and footer information
            items: List of quotation items
            file_path: Path to save the Excel file
        """
        # Load the template file
        template_path = '../data/quotation_template.xlsx'
        self.wb = openpyxl.load_workbook(template_path)
        self.ws = self.wb.active
        
        # Store original merged cells info
        self.original_merged_ranges = list(self.ws.merged_cells.ranges)
        
        # Define styles for new content
        normal_font = Font(name='Angsana New', size=14)
        bold_font = Font(name='Angsana New', size=14, bold=True)
        title_font = Font(name='Angsana New', size=14, bold=True, underline='single')
        center_alignment = Alignment(horizontal='center', vertical='bottom')
        left_alignment = Alignment(horizontal='left', vertical='bottom')
        right_alignment = Alignment(horizontal='right', vertical='bottom')
        
        # === STEP 1: CALCULATE EXACT ROWS NEEDED ===
        template_items_start = 14  # Items start at row 14
        template_items_end = 28    # Items area ends at row 28 (footer starts at 29)
        template_available_rows = template_items_end - template_items_start + 1  # 15 rows available
        
        # Calculate exact rows needed (one row per item, no headers)
        total_rows_needed = len(items)
        
        # Calculate how many extra rows to insert
        rows_to_insert = max(0, total_rows_needed - template_available_rows)
        
        # === STEP 2: INSERT ROWS WITH PROPER FORMATTING ===
        if rows_to_insert > 0:
            self._expand_table_for_items(rows_to_insert)
        
        # === POPULATE HEADER SECTION ===
        # Populate the template fields with quotation data
        
        # TO / Company / Tel (template already has labels, just fill the data)
        self._safe_set_cell_value('D5', quote_data.get('to', ''), normal_font)
        self._safe_set_cell_value('D6', quote_data.get('company', ''), normal_font)
        self._safe_set_cell_value('D8', quote_data.get('tel', ''), normal_font)
        self._safe_set_cell_value('H8', quote_data.get('fax', ''), normal_font)
        
        # Quote No / Date / Project (template already has labels)
        self._safe_set_cell_value('N5', quote_data.get('quote_no', ''), normal_font)
        self._safe_set_cell_value('N6', quote_data.get('date', datetime.now().strftime('%Y-%m-%d')), normal_font)
        self._safe_set_cell_value('N7', quote_data.get('project', ''), normal_font)
        
        # === STEP 3: POPULATE PRODUCT INFORMATION ===
        current_row = 14
        item_no = 1
        sub_total = 0

        for item in items:
            if item.get('is_title', False):
                # Title row - no ID number, show title in B column with bold, underline, and left alignment
                self._safe_set_cell_value(f'A{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'B{current_row}', item.get('title', ''), title_font, left_alignment)
                
                # Clear other columns for title
                self._safe_set_cell_value(f'G{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'H{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'I{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'J{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'K{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'M{current_row}', '', normal_font, right_alignment)
                self._safe_set_cell_value(f'O{current_row}', '', normal_font, center_alignment)
                self._safe_set_cell_value(f'Q{current_row}', '', normal_font, right_alignment)
            else:
                # Regular product row
                # NO. in column A (set as integer to avoid "Number Stored as Text" warning)
                self._safe_set_cell_value(f'A{current_row}', item_no, normal_font, center_alignment)
                
                # MODEL in column B (NO MERGING!)
                self._safe_set_cell_value(f'B{current_row}', item.get('product_code', ''), normal_font, left_alignment)
                
                # FINISHING in column F - Thai text with bottom right alignment
                finish = item.get('finish', '')
                thai_finish = self.get_thai_finishing(finish)
                bottom_right_alignment = Alignment(horizontal='right', vertical='bottom')
                self._safe_set_cell_value(f'F{current_row}', thai_finish, normal_font, bottom_right_alignment)
            
                # Parse size - G(Height) H(x) I(Width) J(Unit) - only for regular items
                size = item.get('size', '')
                
                if 'mm' in size:
                    parts = size.replace('mm', '').split('x')
                    if len(parts) == 2:
                        width_mm = float(parts[0].strip())
                        height_mm = float(parts[1].strip())
                        self._safe_set_cell_value(f'G{current_row}', width_mm, normal_font, center_alignment)
                        self._safe_set_cell_value(f'H{current_row}', 'x', normal_font, center_alignment)
                        self._safe_set_cell_value(f'I{current_row}', height_mm, normal_font, center_alignment)
                        # Unit in column J for mm
                        self._safe_set_cell_value(f'J{current_row}', 'mm', normal_font, center_alignment)
                else:
                    parts = size.replace('"', '').split('x')
                    if len(parts) == 2:
                        width_in = float(parts[0].strip())
                        height_in = float(parts[1].strip())
                        self._safe_set_cell_value(f'G{current_row}', width_in, normal_font, center_alignment)
                        self._safe_set_cell_value(f'H{current_row}', 'x', normal_font, center_alignment)
                        self._safe_set_cell_value(f'I{current_row}', height_in, normal_font, center_alignment)
                        # Unit in column J for inches
                        self._safe_set_cell_value(f'J{current_row}', 'in', normal_font, center_alignment)
                
                # QTY in column K
                quantity = int(item.get('quantity', 1))
                self._safe_set_cell_value(f'K{current_row}', quantity, normal_font, center_alignment)
                
                # UNIT PRICE in column M (changed from L)
                unit_price = float(item.get('unit_price', 0))
                self._safe_set_cell_value(f'M{current_row}', unit_price, normal_font, right_alignment)
                
                # Discount in column O (unchanged)
                discount = item.get('discount', 0)
                if discount > 0:
                    # Set as decimal number (0.1 for 10%) and apply percentage format
                    cell = self.ws[f'O{current_row}']
                    cell.value = discount
                    cell.font = normal_font
                    cell.alignment = center_alignment
                    cell.number_format = '0%'  # Format as percentage
                    discounted_price = unit_price * (1 - discount)
                else:
                    self._safe_set_cell_value(f'O{current_row}', '', normal_font, center_alignment)
                    discounted_price = unit_price
                
                # AMOUNT in column Q (changed from P)
                item_total = discounted_price * quantity
                self._safe_set_cell_value(f'Q{current_row}', item_total, normal_font, right_alignment)
                sub_total += item_total
                
                item_no += 1
            
            current_row += 1
        
        # === POPULATE FOOTER SECTION ===
        # Calculate totals and populate template footer fields
        # Footer starts at row 29 in template, but shifted down by inserted rows
        footer_start_row = 29 + rows_to_insert
        
        # Sub total
        self._safe_set_cell_value(f'Q{footer_start_row}', sub_total, bold_font)
        
        # VAT
        vat = sub_total * 0.07
        self._safe_set_cell_value(f'Q{footer_start_row + 1}', vat, bold_font)
        
        # Grand total
        grand_total = sub_total + vat
        self._safe_set_cell_value(f'A{footer_start_row + 2}', self.thai_baht_text(grand_total), normal_font)
        self._safe_set_cell_value(f'Q{footer_start_row + 2}', grand_total, 
                                 Font(name='Angsana New', size=14, bold=True))
        
        # Footer information - populate template fields
        footer_row = footer_start_row + 4
        self._safe_set_cell_value(f'D{footer_row}', quote_data.get('remarks', '1. ยืนราคา 60 วัน'), normal_font)
        self._safe_set_cell_value(f'M{footer_row}', quote_data.get('quoted_by_name', ''), normal_font, center_alignment)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('payment_term', 'เครดิต 30 วัน'), normal_font)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('delivery_place', ''), normal_font)
        self._safe_set_cell_value(f'O{footer_row}', quote_data.get('purchased_by', ''), normal_font, center_alignment)
        
        footer_row += 1
        self._safe_set_cell_value(f'F{footer_row}', quote_data.get('delivery_date', ''), normal_font)
        
        # Save the workbook
        self.wb.save(file_path)
        return True
    
    def _expand_table_for_items(self, rows_to_insert):
        """
        Expand the table to accommodate more items by inserting rows with proper formatting
        
        Args:
            rows_to_insert: Number of rows to insert
        """
        if rows_to_insert <= 0:
            return
        
        # Insert rows INSIDE the table, just before the footer section
        insert_position = 28  # Insert before the footer section
        
        # === STEP 1: Handle merged cells in footer section ===
        # Store merged cell ranges that are at or below the insert position
        footer_merged_cells = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # Check if this merged range starts at or after the insert position
            if merged_range.min_row >= insert_position:
                # Store the range information
                footer_merged_cells.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
                # Unmerge this range
                self.ws.unmerge_cells(str(merged_range))
        
        # === STEP 2: Insert the rows ===
        reference_row = 27  # Reference row to copy formatting from
        self.ws.insert_rows(insert_position, rows_to_insert)
        
        # === STEP 3: Copy formatting to newly inserted rows ===
        for i in range(rows_to_insert):
            target_row = insert_position + i
            
            # Copy formatting for each column
            for col in range(1, 19):  # A to Q (columns 1-17, extended range)
                source_cell = self.ws.cell(row=reference_row, column=col)
                target_cell = self.ws.cell(row=target_row, column=col)
                
                # Copy font
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color
                    )
                
                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )
                
                # Copy border
                if source_cell.border:
                    target_cell.border = Border(
                        left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
                        right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
                        top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
                        bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                    )
                
                # Copy fill
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
                
                # Copy number format (important for unit prices)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
        
        # === STEP 4: Re-merge cells in footer at new positions ===
        for merged_info in footer_merged_cells:
            # Calculate new row positions (shifted down by rows_to_insert)
            new_min_row = merged_info['min_row'] + rows_to_insert
            new_max_row = merged_info['max_row'] + rows_to_insert
            
            # Reconstruct the range string
            start_cell = f"{get_column_letter(merged_info['min_col'])}{new_min_row}"
            end_cell = f"{get_column_letter(merged_info['max_col'])}{new_max_row}"
            range_string = f"{start_cell}:{end_cell}"
            
            # Re-merge the cells at the new position
            try:
                self.ws.merge_cells(range_string)
            except Exception as e:
                print(f"Warning: Could not re-merge cells {range_string}: {e}")

    def _safe_set_cell_value(self, cell_ref, value, font=None, alignment=None):
        """Safely set cell value, handling merged cells"""
        try:
            cell = self.ws[cell_ref]
            
            # Check if this cell is part of a merged range
            if self._is_merged_cell(cell):
                # Find the top-left cell of the merged range
                for merged_range in self.ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        # Get the top-left cell of the merged range
                        top_left = merged_range.min_row, merged_range.min_col
                        top_left_cell = self.ws.cell(row=top_left[0], column=top_left[1])
                        top_left_cell.value = value
                        if font:
                            top_left_cell.font = font
                        if alignment:
                            top_left_cell.alignment = alignment
                        break
            else:
                # Regular cell, set value directly
                cell.value = value
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                    
        except Exception as e:
            print(f"Warning: Could not set value for {cell_ref}: {e}")
    
    def _safe_merge_cells(self, range_string):
        """Safely merge cells, checking for conflicts"""
        try:
            # Check if any cell in the range is already merged
            start_col, start_row, end_col, end_row = openpyxl.utils.range_boundaries(range_string)
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell_ref = openpyxl.utils.get_column_letter(col) + str(row)
                    for merged_range in self.ws.merged_cells.ranges:
                        if cell_ref in merged_range:
                            # This cell is already merged, skip merging
                            return False
            
            self.ws.merge_cells(range_string)
            return True
        except Exception as e:
            print(f"Warning: Could not merge cells {range_string}: {e}")
            return False
    
    def _is_merged_cell(self, cell):
        """Check if a cell is part of a merged range"""
        try:
            cell_ref = cell.coordinate
            for merged_range in self.ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    return True
            return False
        except:
            return False
    
    def _copy_row_formatting(self, source_row, target_row):
        """Copy all formatting from source row to target row"""
        for col in range(1, 19):  # A to P
            source_cell = self.ws.cell(row=source_row, column=col)
            target_cell = self.ws.cell(row=target_row, column=col)
            
            # Copy font
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            # Copy border
            if source_cell.border:
                target_cell.border = Border(
                    left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
                    right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
                    top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
                    bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                )
            
            # Copy fill
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # Copy number format (important for unit prices)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format