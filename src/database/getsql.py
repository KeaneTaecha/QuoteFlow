"""
AUTOMATIC EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION
===============================================================
This script automatically reads your entire Excel file and converts
ALL tables (even multiple tables per sheet) to SQLite database.

Usage:
    python convert_excel_to_sqlite.py

Output:
    Creates 'prices.db' with all your price data
"""

import sqlite3
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Optional, Dict, Set
import re
from handlers.other_handler import OtherTableHandler
from handlers.default_handler import DefaultTableHandler
from table_models import TableLocation


class ExcelToSQLiteConverter:
    """Automatically convert Excel price list to SQLite database with multi-table detection"""
    
    def __init__(self, excel_file, db_file='prices.db'):
        self.excel_file = excel_file
        self.db_file = db_file
        self.conn = None
        self.other_handler = OtherTableHandler(self.is_inch_value)
        self.default_handler = DefaultTableHandler(self.is_inch_value)
        
        # Will read from Header sheet
        self.header_data = []
        
        # Statistics
        self.stats = {
            'total_sheets': 0,
            'processed_sheets': 0,
            'skipped_sheets': 0,
            'total_prices': 0,
            'total_products': 0,
            'total_tables': 0,
            'errors': []
        }
    
    def create_database(self):
        """Create database structure"""
        print("Creating database structure...")
        
        self.conn = sqlite3.connect(self.db_file)
        cursor = self.conn.cursor()
        
        # Products table
        cursor.execute('''
            CREATE TABLE products (
                product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_id INTEGER NOT NULL,
                model TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                anodized_multiplier REAL,
                powder_coated_multiplier REAL,
                other_paint_multiplier REAL,
                UNIQUE(table_id, model)
            )
        ''')
        
        # Prices table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS prices (
                price_id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_id INTEGER NOT NULL,
                width INTEGER,
                height INTEGER,
                normal_price REAL,
                price_with_damper REAL,
                FOREIGN KEY (table_id) REFERENCES products(table_id) ON UPDATE CASCADE,
                UNIQUE(table_id, width, height)
            )
        ''')
        
        # Indexes for fast lookups
        cursor.execute('''
            CREATE INDEX idx_price_lookup 
            ON prices(table_id, width, height)
        ''')
        
        cursor.execute('''
            CREATE INDEX idx_model_lookup 
            ON products(model)
        ''')
        
        self.conn.commit()
        print("✓ Database structure created\n")
    
    def is_inch_value(self, value) -> Optional[int]:
        """Check if a cell value is an inch measurement (e.g., '4"', '6"')"""
        if value and isinstance(value, str) and '"' in str(value):
            try:
                # Extract inch value from string like "4"" -> 4
                width_str = str(value).replace('"', '').strip()
                return int(width_str)
            except ValueError:
                return None
        return None
    



    def detect_table_at_position(self, sheet, start_row: int, start_col: int, 
                            processed_areas: Set[Tuple[int, int]]) -> Optional[TableLocation]:
        """Detect a table at a specific position and return its boundaries"""
        
        # Check if this area was already processed
        if (start_row, start_col) in processed_areas:
            return None
        
        # Try to find a default/standard table first
        default_table = self.default_handler.get_default_table_bounderies(sheet, start_row, start_col)
        if default_table:
            print(f"        Debug: Found default table at ({start_row}, {start_col})")
            return default_table
        
        # If no default table found, try to detect other table
        other_table = self.other_handler.get_other_table_bounderies(sheet, start_row, start_col)
        if other_table:
            print(f"        Debug: Found other table at ({start_row}, {start_col})")
            return other_table
        
        return None
    
    def detect_all_tables(self, sheet, expected_count: Optional[int] = None) -> List[TableLocation]:
        """Detect all tables in a sheet using pattern recognition"""
        tables = []
        processed_areas = set()
        max_search_row = min(200, sheet.max_row)  # Limit search depth
        max_search_col = min(100, sheet.max_column)  # Limit search width
        
        # Unified detection for all sheets - no special handling based on sheet name
        for row in range(1, max_search_row + 1):
            for col in range(1, max_search_col + 1):
                # Skip if this cell is already part of a processed table
                if any((row >= t.start_row and row <= t.end_row and 
                       col >= t.start_col and col <= t.end_col) for t in tables):
                    continue
                
                # Look for any content in the cell
                cell_value = sheet.cell(row, col).value
                
                # If there's something in the cell, try to find table boundaries
                if cell_value is not None and str(cell_value).strip():
                    # Count values with disruption reset logic
                    values_found = 1  # Start with current cell
                    consecutive_disruptions = 0
                    
                    # Check cells below for inch values
                    for offset in range(1, 7):  # Start from offset 1, not 2
                        if row + offset > max_search_row:  # Bounds checking
                            break
                            
                        below_cell = sheet.cell(row + offset, col)
                        
                        if self.is_inch_value(below_cell.value):
                            # Found an inch value - reset disruption count and increment values
                            values_found += 1
                            consecutive_disruptions = 0
                            
                            # Check if we found enough inch values to consider this a table
                            if values_found >= 3 and consecutive_disruptions <= 1:
                                # Potential table found, try to determine its boundaries
                                table = self.detect_table_at_position(sheet, row, col, processed_areas)
                                if table:
                                    tables.append(table)
                                    # Mark this area as processed
                                    for r in range(table.start_row, table.end_row + 1):
                                        for c in range(table.start_col, table.end_col + 1):
                                            processed_areas.add((r, c))
                                    
                                    # Stop searching if we've found the expected number of tables
                                    if expected_count is not None and len(tables) >= expected_count:
                                        break
                                # Exit the loop once we've found and processed a table
                                break
                        else:
                            # Found disruption - increment consecutive disruption count
                            consecutive_disruptions += 1
                            # Exit if we have too many consecutive disruptions
                            if consecutive_disruptions > 1:
                                break
            
            # Break outer loop if we've found the expected number of tables
            if expected_count is not None and len(tables) >= expected_count:
                break
        
        # Sort tables by position (left to right, then top to bottom)
        tables.sort(key=lambda t: (t.start_row, t.start_col))
        
        return tables
    
    

    def extract_and_store_prices(self, sheet, table_loc: TableLocation, table_id: int) -> int:
        """Extract price data from a table and store in database"""
        if table_loc.table_type == "other":
            return self.other_handler.extract_other_table_prices(sheet, table_loc, table_id, self.conn)
        
        # Extract from default/standard table
        return self.default_handler.extract_default_table_prices(sheet, table_loc, table_id, self.conn)
    
    def read_header_sheet(self, wb):
        """Read the Header sheet to get table metadata"""
        print("Reading Header sheet...")
        
        if 'Header' not in wb.sheetnames:
            print("❌ Error: 'Header' sheet not found in Excel file!")
            return False
        
        header_sheet = wb['Header']
        self.header_data = []
        
        # Group entries by sheet name
        sheet_entries = {}
        
        # Read header data
        for row in header_sheet.iter_rows(min_row=2, values_only=True):
            table_id, sheet_name, model = row[0], row[1], row[2]
            anodized_multiplier, powder_coated_multiplier, other_paint_multiplier = row[3], row[4], row[5]
            
            if table_id is None or sheet_name is None or model is None:
                continue
            
            # Parse models
            models = [m.strip() for m in str(model).split(',')]
            
            # Parse multipliers
            anodized_multipliers = []
            if anodized_multiplier is not None and str(anodized_multiplier).strip().lower() != 'none':
                anodized_multipliers = [float(m.strip()) for m in str(anodized_multiplier).split(',')]
            
            powder_coated_multipliers = []
            if powder_coated_multiplier is not None and str(powder_coated_multiplier).strip().lower() != 'none':
                powder_coated_multipliers = [float(m.strip()) for m in str(powder_coated_multiplier).split(',')]
            
            other_paint_multipliers = []
            if other_paint_multiplier is not None and str(other_paint_multiplier).strip().lower() != 'none':
                other_paint_multipliers = [float(m.strip()) for m in str(other_paint_multiplier).split(',')]
            
            # Adjust multiplier lists to match model count
            if len(anodized_multipliers) == 1:
                anodized_multipliers = anodized_multipliers * len(models)
            elif len(anodized_multipliers) > 0 and len(anodized_multipliers) < len(models):
                anodized_multipliers.extend([anodized_multipliers[-1]] * (len(models) - len(anodized_multipliers)))
            
            if len(powder_coated_multipliers) == 1:
                powder_coated_multipliers = powder_coated_multipliers * len(models)
            elif len(powder_coated_multipliers) > 0 and len(powder_coated_multipliers) < len(models):
                powder_coated_multipliers.extend([powder_coated_multipliers[-1]] * (len(models) - len(powder_coated_multipliers)))
            
            if len(other_paint_multipliers) == 1:
                other_paint_multipliers = other_paint_multipliers * len(models)
            elif len(other_paint_multipliers) > 0 and len(other_paint_multipliers) < len(models):
                other_paint_multipliers.extend([other_paint_multipliers[-1]] * (len(models) - len(other_paint_multipliers)))
            
            entry = {
                'table_id': int(table_id),
                'sheet_name': str(sheet_name),
                'models': models,
                'anodized_multipliers': anodized_multipliers,
                'powder_coated_multipliers': powder_coated_multipliers,
                'other_paint_multipliers': other_paint_multipliers
            }
            
            self.header_data.append(entry)
            
            # Group by sheet name for multi-table detection
            if sheet_name not in sheet_entries:
                sheet_entries[sheet_name] = []
            sheet_entries[sheet_name].append(entry)
        
        # Store grouped entries for easier processing
        self.sheet_entries = sheet_entries
        
        print(f"✓ Found {len(self.header_data)} table(s) in Header sheet")
        print(f"✓ Sheets with multiple tables: {[s for s, e in sheet_entries.items() if len(e) > 1]}\n")
        return True
    
    def insert_products(self, table_id, sheet_name, models, anodized_multipliers, powder_coated_multipliers, other_paint_multipliers):
        """Insert product records for a table"""
        cursor = self.conn.cursor()
        
        for i, model in enumerate(models):
            anodized_mult = anodized_multipliers[i] if i < len(anodized_multipliers) else None
            powder_mult = powder_coated_multipliers[i] if i < len(powder_coated_multipliers) else None
            other_paint_mult = other_paint_multipliers[i] if i < len(other_paint_multipliers) else None
            
            cursor.execute('''
                INSERT OR IGNORE INTO products (table_id, model, sheet_name, anodized_multiplier, powder_coated_multiplier, other_paint_multiplier)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (table_id, model, sheet_name, anodized_mult, powder_mult, other_paint_mult))
            self.stats['total_products'] += 1
    
    def extract_tables_from_sheet(self, sheet, sheet_name):
        """Extract and process all tables from a sheet"""
        entries = self.sheet_entries.get(sheet_name, [])
        if not entries:
            return 0
        
        total_prices = 0
        detected_tables = self.detect_all_tables(sheet, len(entries))
        
        print(f"  Processing {len(entries)} table(s) - detected {len(detected_tables)} table(s)")
        
        # Process each detected table
        for i, table_loc in enumerate(detected_tables):
            if i < len(entries):
                entry = entries[i]
                table_type_str = f" ({table_loc.table_type})" if table_loc.table_type != "standard" else ""
                print(f"    Table {i+1} (ID: {entry['table_id']}, Models: {', '.join(entry['models'][:2])}...){table_type_str}")
                print(f"      Location: Row {table_loc.start_row}, Col {table_loc.start_col}")
                
                # Insert products
                self.insert_products(
                    entry['table_id'], sheet_name, entry['models'],
                    entry['anodized_multipliers'], entry['powder_coated_multipliers'], entry['other_paint_multipliers']
                )
                
                # Process the table
                price_count = self.extract_and_store_prices(sheet, table_loc, entry['table_id'])
                total_prices += price_count
                self.stats['total_tables'] += 1
                print(f"      ✓ {price_count} prices imported")
            else:
                print(f"    ⚠ Extra table detected at ({table_loc.start_row},{table_loc.start_col}) - no Header entry")
        
        # Warn about missing tables
        if len(entries) > len(detected_tables):
            print(f"    ⚠ Warning: Expected {len(entries)} tables but only found {len(detected_tables)}")
            for j in range(len(detected_tables), len(entries)):
                print(f"      Missing: Table ID {entries[j]['table_id']}")
        
        return total_prices
    
    def convert(self):
        """Main conversion process"""
        print("="*70)
        print("AUTOMATIC EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION")
        print("="*70)
        print(f"\nSource file: {self.excel_file}")
        print(f"Output database: {self.db_file}")
        print()
        
        # Check if Excel file exists
        if not Path(self.excel_file).exists():
            print(f"❌ Error: File '{self.excel_file}' not found!")
            return False
        
        # Create database
        self.create_database()
        
        # Load Excel file
        print(f"Loading Excel file...")
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True, data_only=True)
            print(f"✓ Loaded {len(wb.sheetnames)} sheets\n")
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False
        
        self.stats['total_sheets'] = len(wb.sheetnames)
        
        # Read Header sheet
        if not self.read_header_sheet(wb):
            return False
        
        if not self.header_data:
            print("❌ Error: No table information found in Header sheet!")
            return False
        
        # Process sheets
        print("Processing sheets:")
        print("-" * 70)
        
        processed_sheets = set()
        
        for sheet_name in self.sheet_entries.keys():
            if sheet_name in processed_sheets:
                continue
            
            if sheet_name not in wb.sheetnames:
                print(f"❌ Warning: Sheet '{sheet_name}' not found in Excel file!")
                self.stats['skipped_sheets'] += 1
                continue
            
            print(f"\nProcessing: {sheet_name}")
            
            try:
                sheet = wb[sheet_name]
                price_count = self.extract_tables_from_sheet(sheet, sheet_name)
                
                if price_count > 0:
                    print(f"  Total: {price_count} prices from this sheet")
                    self.stats['processed_sheets'] += 1
                    self.stats['total_prices'] += price_count
                else:
                    print(f"  ⚠ No valid prices found in sheet")
                    self.stats['skipped_sheets'] += 1
                
                processed_sheets.add(sheet_name)
                    
            except Exception as e:
                print(f"  ❌ Error: {str(e)}")
                self.stats['errors'].append(f"{sheet_name}: {str(e)}")
                self.stats['skipped_sheets'] += 1
        
        # Commit all changes
        self.conn.commit()
        self.conn.close()
        
        # Print summary
        self.print_summary()
        
        return True
    
    def print_summary(self):
        """Print conversion summary"""
        print("\n" + "="*70)
        print("CONVERSION COMPLETE!")
        print("="*70)
        print(f"\nTotal sheets in Excel:     {self.stats['total_sheets']}")
        print(f"Sheets processed:          {self.stats['processed_sheets']}")
        print(f"Sheets skipped:            {self.stats['skipped_sheets']}")
        print(f"Total tables found:        {self.stats['total_tables']}")
        print(f"Total products created:    {self.stats['total_products']}")
        print(f"Total prices inserted:     {self.stats['total_prices']:,}")
        
        if self.stats['errors']:
            print(f"\nErrors encountered:        {len(self.stats['errors'])}")
            for error in self.stats['errors'][:5]:
                print(f"  - {error}")
        
        print(f"\n✓ Database saved to: {self.db_file}")
        print(f"✓ File size: {Path(self.db_file).stat().st_size / 1024:.1f} KB")
        print("\nYou can now use this database in your PyQt application!")


# =============================================================================
# MAIN PROGRAM - RUN THIS!
# =============================================================================

if __name__ == "__main__":
    # Configuration
    EXCEL_FILE = '../../data/price_list_modified.xlsx'  # Your Excel file name
    DATABASE_FILE = '../../prices.db'               # Output database name
    
    print("\n")
    print("╔════════════════════════════════════════════════════════════════════╗")
    print("║       EXCEL TO SQLITE CONVERTER WITH MULTI-TABLE DETECTION         ║")
    print("║       Automatically detects and processes multiple tables!         ║")
    print("╚════════════════════════════════════════════════════════════════════╝")
    print("\n")
    
    # Create converter
    converter = ExcelToSQLiteConverter(EXCEL_FILE, DATABASE_FILE)
    
    # Run conversion
    success = converter.convert()
    
    if success:
        print("\n✓ Conversion completed successfully!")
    else:
        print("\n❌ Conversion failed. Please check the errors above.")